from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def generate_sample_sales(output_path: Path) -> None:
    data = [
        {"Product": "Laptop Pro 14", "Unit_Price": 1499.0, "Quantity_Sold": 18},
        {"Product": "Wireless Mouse", "Unit_Price": 29.5, "Quantity_Sold": 240},
        {"Product": "USB-C Dock", "Unit_Price": 189.0, "Quantity_Sold": 65},
        {"Product": "4K Monitor", "Unit_Price": 399.0, "Quantity_Sold": 42},
        {"Product": "Noise-Cancel Headset", "Unit_Price": 219.0, "Quantity_Sold": 75},
        {"Product": "Ergonomic Keyboard", "Unit_Price": 119.0, "Quantity_Sold": 110},
    ]
    df = pd.DataFrame(data, columns=["Product", "Unit_Price", "Quantity_Sold"])
    df.to_excel(output_path, index=False)


def format_sales_report(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="C7C7C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    column_map = {cell.value: cell.column for cell in worksheet[1]}
    currency_format = '"$"#,##0.00'
    integer_format = "#,##0"

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        unit_price_col = column_map.get("Unit_Price")
        quantity_col = column_map.get("Quantity_Sold")
        total_col = column_map.get("Total_Revenue")

        if unit_price_col:
            worksheet.cell(row=row, column=unit_price_col).number_format = currency_format
        if total_col:
            worksheet.cell(row=row, column=total_col).number_format = currency_format
        if quantity_col:
            worksheet.cell(row=row, column=quantity_col).number_format = integer_format

    for col_idx in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in worksheet[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = max(12, min(max_length + 2, 40))

    table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    table = Table(displayName="SalesSummary", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    workbook.save(output_path)


def build_sales_summary(input_path: Path, output_path: Path) -> None:
    df = pd.read_excel(input_path)
    df["Total_Revenue"] = df["Unit_Price"] * df["Quantity_Sold"]
    df.to_excel(output_path, index=False)
    format_sales_report(output_path)


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    input_path = base_dir / "sales_data.xlsx"
    output_path = base_dir / "sales_summary.xlsx"

    generate_sample_sales(input_path)
    build_sales_summary(input_path, output_path)

    print(f"Sales summary report created: {output_path}")


if __name__ == "__main__":
    main()
